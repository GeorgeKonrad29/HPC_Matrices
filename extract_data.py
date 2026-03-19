#!/usr/bin/env python3

import os
import re
import csv
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

def extract_data(file_path):
    """Extrae tamaño y tiempo de un archivo de resultados"""
    results = []
    
    try:
        with open(file_path, 'r') as f:
            content = f.read()
    except Exception as e:
        print(f"Error leyendo {file_path}: {e}")
        return results
    
    # Buscar líneas con tamaño
    size_patterns = [
        r'Tamaño:\s*(\d+)x(\d+)',
        r'Tamaño de las matrices:\s*(\d+)x(\d+)'
    ]
    
    # Buscar líneas con tiempo
    time_patterns = [
        r'Tiempo:\s*([\d.]+)\s*segundos',
        r'Tiempo de multiplicación:\s*([\d.]+)\s*segundos'
    ]
    
    lines = content.split('\n')
    execution_num = 0
    current_size = None
    
    for line in lines:
        # Buscar tamaño
        for pattern in size_patterns:
            size_match = re.search(pattern, line)
            if size_match:
                current_size = f"{size_match.group(1)}x{size_match.group(2)}"
                break
        
        # Buscar tiempo
        for pattern in time_patterns:
            time_match = re.search(pattern, line)
            if time_match:
                execution_num += 1
                tiempo = float(time_match.group(1))
                if current_size:
                    results.append({
                        'num': execution_num,
                        'size': current_size,
                        'time': tiempo
                    })
                break
    
    return results

def save_to_csv(all_data):
    """Guarda los datos en CSV"""
    output_file = "resultados.csv"
    
    try:
        with open(output_file, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['Archivo', 'Ejecución', 'Tamaño', 'Tiempo (s)'])
            
            for file_name, results in all_data.items():
                for result in results:
                    writer.writerow([
                        file_name,
                        result['num'],
                        result['size'],
                        f"{result['time']:.4f}"
                    ])
        
        print(f"\n✓ Datos guardados en: {output_file}")
        return True
    except Exception as e:
        print(f"Error guardando CSV: {e}")
        return False

def save_to_excel(all_data):
    """Guarda los datos en Excel"""
    if not OPENPYXL_AVAILABLE:
        return False
    
    output_file = "resultados.xlsx"
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Resultados"
        
        # Estilos
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        center = Alignment(horizontal="center", vertical="center")
        
        # Encabezados
        headers = ['Archivo', 'Ejecución', 'Tamaño', 'Tiempo (s)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
        
        # Datos
        row = 2
        for file_name, results in sorted(all_data.items()):
            for result in results:
                ws.cell(row=row, column=1, value=file_name)
                ws.cell(row=row, column=2, value=result['num'])
                ws.cell(row=row, column=3, value=result['size'])
                ws.cell(row=row, column=4, value=f"{result['time']:.4f}")
                row += 1
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        
        wb.save(output_file)
        print(f"✓ Datos guardados en: {output_file}")
        return True
    except Exception as e:
        print(f"Error guardando Excel: {e}")
        return False

def main():
    # Directorio actual
    work_dir = Path.cwd()
    
    # Archivos a procesar
    doc_files = sorted(work_dir.glob('*.doc'))
    
    if not doc_files:
        print("No se encontraron archivos .doc")
        return
    
    print("=" * 70)
    print("EXTRACCIÓN DE DATOS - MULTIPLICACIÓN DE MATRICES")
    print("=" * 70)
    print()
    
    # Almacenar todos los datos
    all_data = {}
    
    for file_path in doc_files:
        file_name = file_path.name
        print(f"\n{'='*70}")
        print(f"Archivo: {file_name}")
        print(f"{'='*70}")
        print(f"{'Ejecución':<12} {'Tamaño':<15} {'Tiempo (s)':<15}")
        print("-" * 70)
        
        results = extract_data(file_path)
        all_data[file_name] = results
        
        if not results:
            print("No se extrajeron datos")
            continue
        
        for result in results:
            print(f"{result['num']:<12} {result['size']:<15} {result['time']:<15.4f}")
        
        print(f"\nTotal ejecuciones: {len(results)}")
    
    # Guardar datos
    print("\n" + "=" * 70)
    print("GUARDANDO DATOS...")
    print("=" * 70)
    
    save_to_csv(all_data)
    if OPENPYXL_AVAILABLE:
        save_to_excel(all_data)
    else:
        print("⚠ openpyxl no está instalado. Solo se guardó CSV.")
        print("  Instala con: pip install openpyxl")

if __name__ == "__main__":
    main()
